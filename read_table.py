from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
wb = load_workbook(filename = 'sample1.xlsx')
ws = wb.active
# cell = 2:2 i.e B:2 tablecorner = topleft,bottomright, topright, bottomleft
def checkTableExist(table,tableDict, cell,tablecorner):
    #print(topleft)
    #print(table in tableDict)
    
    
    for x in tableDict:
       
        if tableDict[x].get(tablecorner) == cell:
            return True
        

        
        
            
    #print("Table Not Found") 
    return False

def isTopLeftCornerOfTable(ws,col,row):
    if (col > 1 and row > 1):
        topCell = NumToAlpha(col)+str(row-1)
        leftCell= NumToAlpha(col-1)+str(row)
        diaCell = NumToAlpha(col-1)+str(row-1)
        #print(topCell,leftCell,diaCell)
        topAdjCellBorders = getCellBorders(ws,topCell)
        leftAdjCellBorders = getCellBorders(ws,leftCell)
        diaAdjCellBorders = getCellBorders(ws,diaCell)
        #print(topAdjCellBorders,leftAdjCellBorders,diaAdjCellBorders)
        if ('L' not in topAdjCellBorders and 'R' not in topAdjCellBorders and 
            'B' not in leftAdjCellBorders and 'T' not in leftAdjCellBorders and 
            'B' not in diaAdjCellBorders and 'R' not in diaAdjCellBorders
            ):
            #print("Top Left")
            return True
        else:
            if topAdjCellBorders == '' and leftAdjCellBorders == '' and diaAdjCellBorders == '' :
                #print("Top Left")
                return True
            else:
                #print("No")
                return False
    # first cell
    if (col == 1 and row == 1):
        return True
    
    if (col > 1 and row == 1):
        leftCell= NumToAlpha(col-1)+str(row)
        leftAdjCellBorders = getCellBorders(ws,leftCell)
        if ('B' not in leftAdjCellBorders and 'T' not in leftAdjCellBorders):
            return True
        else:
            if (leftAdjCellBorders == ''):
                return True
            else:
                return False
    if (col == 1 and row > 1):
        topCell= NumToAlpha(col)+str(row-1)
        topAdjCellBorders = getCellBorders(ws,topCell)
        if ('L' not in topAdjCellBorders and 'R' not in topAdjCellBorders):
            return True
        else:
            if (topAdjCellBorders == ''):
                return True
            else:
                return False
        

def isTopRightCornerOfTable(ws,col,row):
    if (col > 1 and row > 1):
        topCell = NumToAlpha(col)+str(row-1)
        rightCell= NumToAlpha(col+1)+str(row)
        diaCell = NumToAlpha(col+1)+str(row-1)
        #print(topCell,leftCell,diaCell)
        topAdjCellBorders = getCellBorders(ws,topCell)
        rightAdjCellBorders = getCellBorders(ws,rightCell)
        diaAdjCellBorders = getCellBorders(ws,diaCell)
        #print(topAdjCellBorders,leftAdjCellBorders,diaAdjCellBorders)
        if ('L' not in topAdjCellBorders and 'R' not in topAdjCellBorders and 
            'B' not in rightAdjCellBorders and 'T' not in rightAdjCellBorders and 
            'B' not in diaAdjCellBorders and 'L' not in diaAdjCellBorders
            ):
            #print("Top Right")
            return True
        else:
            if topAdjCellBorders == '' and rightAdjCellBorders == '' and diaAdjCellBorders == '' :
                #print("Top Right")
                return True
            else:
                #print("No")
                return False
            
    
    
    if (col >= 1 and row == 1):
        rightCell= NumToAlpha(col+1)+str(row)
        rightAdjCellBorders = getCellBorders(ws,rightCell)
        if ('B' not in rightAdjCellBorders and 'T' not in rightAdjCellBorders):
            return True
        else:
            if (rightAdjCellBorders == ''):
                return True
            else:
                return False
            


def isBottomLeftCornerOfTable(ws,col,row):
    if (col > 1 and row > 1):
        bottomCell = NumToAlpha(col)+str(row+1)
        leftCell= NumToAlpha(col-1)+str(row)
        diaCell = NumToAlpha(col-1)+str(row+1)
        #print(topCell,leftCell,diaCell)
        leftAdjCellBorders = getCellBorders(ws,leftCell)
        bottomAdjCellBorders = getCellBorders(ws,bottomCell)
        diaAdjCellBorders = getCellBorders(ws,diaCell)
        #print(topAdjCellBorders,leftAdjCellBorders,diaAdjCellBorders)
        if ('L' not in bottomAdjCellBorders and 'R' not in bottomAdjCellBorders and 
            'B' not in leftAdjCellBorders and 'T' not in leftAdjCellBorders and 
            'T' not in diaAdjCellBorders and 'R' not in diaAdjCellBorders
            ):
            #print("Top Right")
            return True
        else:
            if bottomAdjCellBorders == '' and leftAdjCellBorders == '' and diaAdjCellBorders == '' :
                #print("Top Right")
                return True
            else:
                #print("No")
                return False
    if (col == 1 and row >=1):
        bottomCell = NumToAlpha(col)+str(row+1)
        bottomAdjCellBorders = getCellBorders(ws,bottomCell)
        if ('L' not in bottomAdjCellBorders and 'R' not in bottomAdjCellBorders):
            return True
        else:
            if bottomAdjCellBorders == '':
                return True
            else:
                #print("No")
                return False


def isBottomRightCornerOfTable(ws,col,row):
    if (col >= 1 and row >= 1):
        bottomCell = NumToAlpha(col)+str(row+1)
        rightCell= NumToAlpha(col+1)+str(row)
        diaCell = NumToAlpha(col+1)+str(row+1)
        #print(topCell,leftCell,diaCell)
        rightAdjCellBorders = getCellBorders(ws,rightCell)
        bottomAdjCellBorders = getCellBorders(ws,bottomCell)
        diaAdjCellBorders = getCellBorders(ws,diaCell)
        #print(topAdjCellBorders,leftAdjCellBorders,diaAdjCellBorders)
        if ('L' not in bottomAdjCellBorders and 'R' not in bottomAdjCellBorders and 
            'B' not in rightAdjCellBorders and 'T' not in rightAdjCellBorders and 
            'T' not in diaAdjCellBorders and 'L' not in diaAdjCellBorders
            ):
            #print("Top Right")
            return True
        else:
            if bottomAdjCellBorders == '' and rightAdjCellBorders == '' and diaAdjCellBorders == '' :
                #print("Top Right")
                return True
            else:
                #print("No")
                return False

    

def getCellBorders(ws, cellRef):
    tmp = ws[cellRef].border
    #print(tmp.top)
    brdrs = ''
    if tmp.top is not None:
        if tmp.top.style is not None: brdrs += 'T'
    if tmp.left is not None:
        if tmp.left.style is not None: brdrs += 'L'
    if tmp.right is not None:
        if tmp.right.style is not None: brdrs += 'R'
    if tmp.bottom is not None:
        if tmp.bottom.style is not None: brdrs += 'B'
    return brdrs
def NumToAlpha(col):
    alphaList = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG']
    return alphaList[col-1]
# Iterating through All rows with all columns...

tableDict = {}
for table in range(1,10,1):
    tableTopLeftCornerFound = False
    tableTopRightCornerFound = False
    tableBottomLeftCornerFound = False
    tableBottomRightCornerFound = False
    tableDict["table"+str(table)] = {}
    #print(tableDict)
    tablecount = table
    tablecount = str(tablecount)
    #print(type(tablecount))
    for row in range(1, ws.max_row, 1):
        
        for col in range(1, 33):
            
            if (tableTopLeftCornerFound and tableTopRightCornerFound and tableBottomLeftCornerFound and tableBottomRightCornerFound):
                break
            tmp = NumToAlpha(col)
            rownum = str(row)
            cellRef = str(tmp) + str(row)
            cellBorders = getCellBorders(ws, cellRef)
            
            #print(cellBorders)
            #print(ws[cellRef].value)
            # check if this is top left corder of table
            if ('L' in cellBorders and 'T' in cellBorders and 
                tableTopLeftCornerFound == False ):
                if isTopLeftCornerOfTable(ws,col,row):
                    
                    #print(checkTableExist(tableDict,tmp+":"+rownum))
                    if (checkTableExist("table"+tablecount,tableDict,str(col)+":"+str(row),"topleft") == False):
                        #print (f"Top Left Found for {tmp} {row} table {table}")
                        #tableDict["table"+tablecount]["topleft"]= tmp+":"+rownum
                        tableDict["table"+tablecount]["topleft"]= str(col)+":"+str(row)
                        tableTopLeftCornerFound = True
                
            # check if this is top right corder of table 
            if (tableTopRightCornerFound == False and 
                tableTopLeftCornerFound == True and 
                ('T' in cellBorders) and ('R' in cellBorders)):
                #print(ws[cellRef].value)
                if isTopRightCornerOfTable(ws,col,row):
                    
                    if (checkTableExist("table"+tablecount,tableDict,str(col)+":"+str(row),"topright") == False):
                        #print (f"Top Right Found for {tmp} {row} table {table}")
                        tableDict["table"+tablecount]["topright"]=str(col)+":"+str(row)
                        tableTopRightCornerFound = True

            # check if this is bottom left corder of table
            if (tableBottomLeftCornerFound == False and 
                #tableTopRightCornerFound == True and 
                tableTopLeftCornerFound == True and 
                ('L' in cellBorders) and ('B' in cellBorders)):
                #print(ws[cellRef].value)
                if isBottomLeftCornerOfTable(ws,col,row):
                    
                    if (checkTableExist("table"+tablecount,tableDict,str(col)+":"+str(row),"bottomleft") == False):
                        columntopleft = int(tableDict["table"+tablecount].get("topleft").split(":")[0])
                       
                        if (col == columntopleft):
                            #print (f"Bottom Left Found for {tmp} {row} table {table}")
                            tableDict["table"+tablecount]["bottomleft"]=str(col)+":"+str(row)
                            tableBottomLeftCornerFound = True

            if (tableBottomRightCornerFound == False and
                tableBottomLeftCornerFound == True and 
                tableTopRightCornerFound == True and 
                tableTopLeftCornerFound == True and 
                ('R' in cellBorders) and ('B' in cellBorders)):
                if isBottomRightCornerOfTable(ws,col,row):
                    
                    if (checkTableExist("table"+tablecount,tableDict,str(col)+":"+str(row),"bottomright") == False):
                        columntopright = int(tableDict["table"+tablecount].get("topright").split(":")[0])
                       
                        if (col >= columntopright):
                            
                            #print (f"Bottom Right Found for {tmp} {row} table {table}")
                            tableDict["table"+tablecount]["bottomright"]=str(col)+":"+str(row)
                            tableBottomRightCornerFound = True
        
        if (tableTopLeftCornerFound and tableTopRightCornerFound and tableBottomLeftCornerFound and tableBottomRightCornerFound):
                break

for key, value in tableDict.items():
    if (len(value) > 0):
        print(f"{key}: {value}")