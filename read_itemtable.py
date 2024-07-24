from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import rows_from_range, cols_from_range

def checkInRange(sheet, mergedranges):
    for range in mergedranges:
        if (cell.coordinate in range):
            return True
        
    return False
    

def copy_range(range_str, src, dst):
    row_counter = 0
    print(range_str)
    for row in rows_from_range(range_str):
        row_counter+=1
        col_counter = 0
        for cell in row:
            col_counter+=1
            dst.cell(row=row_counter, column=col_counter).value = src[cell].value
           
            

    return

def isTopLeftCornerOfTable(ws,col,row):
    if (col > 1 and row > 1):
        topCell = NumToAlpha(col)+str(row-1)
        leftCell= NumToAlpha(col-1)+str(row)
        diaCell = NumToAlpha(col-1)+str(row-1)
        #print(topCell,leftCell,diaCell)
        topAdjCellBorders = getCellBorders(ws,topCell)
        leftAdjCellBorders = getCellBorders(ws,leftCell)
        diaAdjCellBorders = getCellBorders(ws,diaCell)
        #print(topAdjCellBorders,leftAdjCellBorders,diaAdjCellBorders)
        if (('L' not in topAdjCellBorders and 'R' not in topAdjCellBorders and 
            'B' not in leftAdjCellBorders and 'T' not in leftAdjCellBorders and 
            'B' not in diaAdjCellBorders and 'R' not in diaAdjCellBorders
            )):
            #print("Top Left")
            return True
        else:
            if topAdjCellBorders == '' and leftAdjCellBorders == '' and diaAdjCellBorders == '' :
                #print("Top Left")
                return True
            else:
                #print("No")
                return False
    # first cell
    if (col == 1 and row == 1):
        return True
    
    if (col > 1 and row == 1):
        leftCell= NumToAlpha(col-1)+str(row)
        leftAdjCellBorders = getCellBorders(ws,leftCell)
        if ('B' not in leftAdjCellBorders and 'T' not in leftAdjCellBorders):
            return True
        else:
            if (leftAdjCellBorders == ''):
                return True
            else:
                return False
    if (col == 1 and row > 1):
        topCell= NumToAlpha(col)+str(row-1)
        topAdjCellBorders = getCellBorders(ws,topCell)
        if ('L' not in topAdjCellBorders and 'R' not in topAdjCellBorders):
            return True
        else:
            if (topAdjCellBorders == ''):
                return True
            else:
                return False
        

def isTopRightCornerOfTable(ws,col,row):
    if (col > 1 and row > 1):
        topCell = NumToAlpha(col)+str(row-1)
        rightCell= NumToAlpha(col+1)+str(row)
        diaCell = NumToAlpha(col+1)+str(row-1)
        #print(topCell,leftCell,diaCell)
        topAdjCellBorders = getCellBorders(ws,topCell)
        rightAdjCellBorders = getCellBorders(ws,rightCell)
        diaAdjCellBorders = getCellBorders(ws,diaCell)
        #print(topAdjCellBorders,leftAdjCellBorders,diaAdjCellBorders)
        if ('L' not in topAdjCellBorders and 'R' not in topAdjCellBorders and 
            'B' not in rightAdjCellBorders and 'T' not in rightAdjCellBorders and 
            'B' not in diaAdjCellBorders and 'L' not in diaAdjCellBorders
            ):
            #print("Top Right")
            return True
        else:
            if topAdjCellBorders == '' and rightAdjCellBorders == '' and diaAdjCellBorders == '' :
                #print("Top Right")
                return True
            else:
                #print("No")
                return False
            
    
    
    if (col >= 1 and row == 1):
        rightCell= NumToAlpha(col+1)+str(row)
        rightAdjCellBorders = getCellBorders(ws,rightCell)
        if ('B' not in rightAdjCellBorders and 'T' not in rightAdjCellBorders):
            return True
        else:
            if (rightAdjCellBorders == ''):
                return True
            else:
                return False
            


def isBottomLeftCornerOfTable(ws,col,row):
    if (col > 1 and row > 1):
        bottomCell = NumToAlpha(col)+str(row+1)
        leftCell= NumToAlpha(col-1)+str(row)
        diaCell = NumToAlpha(col-1)+str(row+1)
        #print(topCell,leftCell,diaCell)
        leftAdjCellBorders = getCellBorders(ws,leftCell)
        bottomAdjCellBorders = getCellBorders(ws,bottomCell)
        diaAdjCellBorders = getCellBorders(ws,diaCell)
        #print(topAdjCellBorders,leftAdjCellBorders,diaAdjCellBorders)
        if ('L' not in bottomAdjCellBorders and 'R' not in bottomAdjCellBorders and 
            'B' not in leftAdjCellBorders and 'T' not in leftAdjCellBorders and 
            'T' not in diaAdjCellBorders and 'R' not in diaAdjCellBorders
            ):
            #print("Top Right")
            return True
        else:
            if bottomAdjCellBorders == '' and leftAdjCellBorders == '' and diaAdjCellBorders == '' :
                #print("Top Right")
                return True
            else:
                #print("No")
                return False
    if (col == 1 and row >=1):
        bottomCell = NumToAlpha(col)+str(row+1)
        bottomAdjCellBorders = getCellBorders(ws,bottomCell)
        if ('L' not in bottomAdjCellBorders and 'R' not in bottomAdjCellBorders):
            return True
        else:
            if bottomAdjCellBorders == '':
                return True
            else:
                #print("No")
                return False


def isBottomRightCornerOfTable(ws,col,row):
    if (col >= 1 and row >= 1):
        bottomCell = NumToAlpha(col)+str(row+1)
        rightCell= NumToAlpha(col+1)+str(row)
        diaCell = NumToAlpha(col+1)+str(row+1)
        #print(topCell,leftCell,diaCell)
        rightAdjCellBorders = getCellBorders(ws,rightCell)
        bottomAdjCellBorders = getCellBorders(ws,bottomCell)
        diaAdjCellBorders = getCellBorders(ws,diaCell)
        #print(topAdjCellBorders,leftAdjCellBorders,diaAdjCellBorders)
        if ('L' not in bottomAdjCellBorders and 'R' not in bottomAdjCellBorders and 
            'B' not in rightAdjCellBorders and 'T' not in rightAdjCellBorders and 
            'T' not in diaAdjCellBorders and 'L' not in diaAdjCellBorders
            ):
            #print("Top Right")
            return True
        else:
            if bottomAdjCellBorders == '' and rightAdjCellBorders == '' and diaAdjCellBorders == '' :
                #print("Top Right")
                return True
            else:
                #print("No")
                return False

    

def getCellBorders(ws, cellRef):
    tmp = ws[cellRef].border
    #print(f"{cellRef} {ws[cellRef].fill.bgColor}")
    #print(tmp)
    #print(tmp.top)
    brdrs = ''
    if tmp.top is not None:
        if tmp.top.style is not None: brdrs += 'T'
    if tmp.left is not None:
        if tmp.left.style is not None: brdrs += 'L'
    if tmp.right is not None:
        if tmp.right.style is not None: brdrs += 'R'
    if tmp.bottom is not None:
        if tmp.bottom.style is not None: brdrs += 'B'
    return brdrs
def NumToAlpha(col):
    alphaList = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI']
    return alphaList[col-1]


def getTableHeaderRow(colored_cell_per_row,words_found_per_row):
    #row_with_max_colored_head = max(zip(colored_cell_per_row.values(), colored_cell_per_row.keys()))[1]
    sorted_dict_by_val = dict(sorted(words_found_per_row.items(), key=lambda item: item[1],reverse=True))
    return list(sorted_dict_by_val.keys())[0]
    print(list(sorted_dict_by_val.keys())[0])
    row_with_max_keyword_found = max(zip(words_found_per_row.values(), words_found_per_row.keys()))[1]
    #print(f"row_with_max_colored_head {row_with_max_colored_head} row_with_max_keyword_found {row_with_max_keyword_found}")
    #if (row_with_max_colored_head == row_with_max_keyword_found):
    return row_with_max_keyword_found
    if (row_with_max_keyword_found != row_with_max_colored_head):
        return row_with_max_keyword_found
    else:
        return -1

filename = "sample2.xlsx"
wb = load_workbook(filename ,data_only=True)
print(wb.sheetnames)
#print(wb.get_sheet_names())

keywords_list = ['code','item','price','qty','quantity','style','size','color']
ws = wb.worksheets[0]
newsheet = wb.create_sheet(wb.sheetnames[0]+"new")
#print(type(ws.merged_cells.ranges))
# variables for table corners
tableTopLeftCornerFound = False
tableTopRightCornerFound = False
tableBottomLeftCornerFound = False
tableBottomRightCornerFound = False
cell_with_border = False
last_row_with_border = -1
tableDict = {"maintable":{}}
keyword_found_count = 0
words_found_per_row = {}
colored_cell_count_per_row = 0
colored_cell_per_row = {}
header_row = 0
for row in ws.iter_rows(min_row=1,max_row=ws.max_row,min_col=1,max_col=35):
    # search in a col values in a rows
    
    for cell in row:
        # for the cell tagname with its bgColor not being 'FFFFFFFF' and '000000000'
        # Assumption headers for any table will be highlighted with some color background
        #print(f"For Cell {cell.coordinate} pattern is {cell.font.bold} and {cell.fill.bgColor.rgb}")
        #print(getMergedCellVal(ws,cell))
        if (cell.fill.tagname == 'patternFill' and 
                cell.fill.bgColor.rgb is not None and 
                cell.fill.bgColor.rgb != 'FFFFFFFF' and 
                #cell.fill.bgColor.rgb != '00000000' and
                #getMergedCellVal(ws,cell) is not None 
                cell.value is not None or 
                (cell.value is None and checkInRange(cell, ws.merged_cells.ranges))
                ):
            colored_cell_count_per_row +=1
            #print(f"For Cell {cell.coordinate} pattern is {cell.fill.tagname}")
            # check if the cell is not None and its value does end with ":"
            if (cell.value is not None and str(cell.value).strip().endswith(":") == False):
               
            
                for xs in keywords_list:
                    if (str(cell.value).lower().find(xs) != -1):
                        keyword_found_count+=1
                        #print(f"{xs} found in {cell.coordinate}")

    words_found_per_row[cell.row] = keyword_found_count
    colored_cell_per_row[cell.row] = colored_cell_count_per_row
    colored_cell_count_per_row = 0      
    keyword_found_count = 0
#print(f"Words Found Per Row {words_found_per_row}") 
#print(f"Colored Cell Found Per Row indicating header row {colored_cell_per_row}") 
header_row = getTableHeaderRow(colored_cell_per_row,words_found_per_row)
print(header_row)
if (header_row != -1):
    # the main table header row
    print(getTableHeaderRow(colored_cell_per_row,words_found_per_row))
    # parse thru the table to fetch rows
    top_left_header = -1
    top_right_header = -1
    for row in ws.iter_rows(min_row=header_row,max_row=ws.max_row,min_col=1,max_col=35):
        if (header_row > 0):
            cell_with_border = False

        for cell in row:
            #if (cell.row==header_row):
            #print(f"{cell.coordinate} has value {cell.value} topleft {top_left_header} tableTopLeftCornerFound {tableTopLeftCornerFound} tableTopRightCornerFound {tableTopRightCornerFound} ")
            tmp = NumToAlpha(cell.column)
            rownum = str(cell.row)
            cellRef = str(tmp) + str(cell.row)    
            if (cell.fill.tagname == 'patternFill' and 
                cell.fill.bgColor.rgb is not None and 
                cell.fill.bgColor.rgb != 'FFFFFFFF' and 
                #cell.fill.bgColor.rgb != '00000000' and 
                #getMergedCellVal(ws,cell) is not None and
                cell.value is not None and 
                cell.row==header_row or 
                (cell.value is None and checkInRange(cell, ws.merged_cells.ranges) and cell.row==header_row)):
                if (top_left_header == -1):
                    top_left_header = cell.column
                    tableDict["maintable"]["topleft"]= NumToAlpha(cell.column)+str(cell.row)
                    tableTopLeftCornerFound = True

            if ((top_left_header > -1) and cell.value is None and 
                checkInRange(cell, ws.merged_cells.ranges) == False and
                getCellBorders(ws, cellRef) == '' and
                cell.row==header_row and 
                tableTopLeftCornerFound == True and
                tableTopRightCornerFound == False):
                top_right_header = cell.column
                #print(cell.coordinate)
                tableDict["maintable"]["topright"]= NumToAlpha(cell.column)+str(cell.row)
                tableTopRightCornerFound = True
                
                last_row_with_border = cell.row

           
            # Find the bottom left and bottom right
            # find row with no cells having borders in the range of table headers
            #print(f"{cell.row} has value topleft {top_left_header} {top_right_header} tableTopLeftCornerFound {tableTopLeftCornerFound} tableTopRightCornerFound {tableTopRightCornerFound} ")
            if (cell.row > header_row and top_left_header > -1 and top_right_header > -1 and cell.column < top_right_header):
            
                
                
                cellBorders = getCellBorders(ws, cellRef)
                #print("Cell Border")
                if (cellBorders != ''):
                    #print("Cell Border")
                    last_row_with_border = cell.row
                    cell_with_border = True
                
            
            

                    



                #if (cell.row > 70):
                    #print(f"{cell.coordinate} has value {cellBorders} topleft {top_left_header} tableTopLeftCornerFound {tableTopLeftCornerFound} tableTopRightCornerFound {tableTopRightCornerFound} ")
                
            

                
        # Check if there is no column found with border
        if (cell_with_border == False and last_row_with_border != header_row):
            
            tableDict["maintable"]["bottomleft"] = NumToAlpha(top_left_header)+str(last_row_with_border)
            tableDict["maintable"]["bottomright"] = NumToAlpha(top_right_header)+str(last_row_with_border)
            copy_range(tableDict["maintable"]["topleft"]+":"+tableDict["maintable"]["bottomright"],ws,newsheet)
            
            wb.save(filename+"_new.xlsx")
            break
           
    print(tableDict)